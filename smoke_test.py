from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from generate_sample import sample_files, write_samples
from src.excel_io import build_result_xlsx
from src.extract import InputFormatError
from src.run import review_files


def main() -> None:
    write_samples()
    result = review_files(sample_files(), api_key="")
    print("files", result.file_count, "ok", result.ok_count, "issues", result.issue_count)
    print(result.rows[["判定", "請求元", "税込金額", "理由"]].to_string(index=False))
    assert result.file_count == 6
    assert result.ok_count == 1
    assert result.issue_count == 5
    assert result.ok_gross == 60500
    assert not result.used_llm
    assert not result.llm_called
    xlsx = build_result_xlsx(result)
    assert len(xlsx) > 1000
    wb = load_workbook(BytesIO(xlsx))
    assert wb.sheetnames == ["転記一覧", "要確認", "判定基準", "処理メモ"]
    assert wb["判定基準"]["A1"].value == "判定の基準"
    meta = {row[0]: row[1] for row in wb["処理メモ"].iter_rows(min_row=2, max_col=2, values_only=True)}
    assert meta["生成AIを呼んだか"] == "いいえ"
    assert meta["生成AIが項目を埋めたか"] == "いいえ"
    issues_ws = wb["要確認"]
    assert issues_ws.max_row == 6  # header + 5
    Path("samples/請求書_転記チェック_確認用.xlsx").write_bytes(xlsx)
    print("ok", len(xlsx), "bytes")
    try:
        review_files([])
        raise SystemExit("should have failed")
    except InputFormatError as exc:
        print("format_error_ok", str(exc)[:40])


if __name__ == "__main__":
    main()
