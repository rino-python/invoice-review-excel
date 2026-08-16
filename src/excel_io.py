"""検収結果を、人が開いてすぐ使える Excel バイト列にする。"""

from __future__ import annotations

import math
import unicodedata
from io import BytesIO

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

from src.colors import HEADER_HEX, STATUS_HEX
from src.run import CLIENT_NAME, ReviewResult, order_columns
from src.validate import judgment_criteria_rows

HEADER_FILL = PatternFill("solid", fgColor=HEADER_HEX)
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATUS_FILLS = {label: PatternFill("solid", fgColor=hex_) for label, hex_ in STATUS_HEX.items()}
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
YEN_COLUMNS = {"税抜金額", "消費税", "税込金額"}


def _style_header(ws, max_col: int | None = None, max_row: int | None = None) -> None:
    last_col = max_col or ws.max_column
    last_row = max_row or ws.max_row
    for col in range(1, last_col + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"


def _display_width(text: str) -> float:
    lines = text.split("\n") if text else [""]
    widest = 0.0
    for line in lines:
        width = 0.0
        for char in line:
            width += 2.0 if unicodedata.east_asian_width(char) in {"W", "F"} else 1.0
        widest = max(widest, width)
    return widest


def _cell_display_text(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    fmt = cell.number_format or "General"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return ""
        if "¥" in fmt or "￥" in fmt:
            return f"¥{int(round(value)):,}"
        if "#,##0" in fmt:
            return f"{int(round(value)):,}"
    return str(value)


def _autosize(ws, min_width: float = 8, max_width: float = 42, header_extra: float = 3.0) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0.0
        for index, cell in enumerate(col):
            width = _display_width(_cell_display_text(cell))
            if index == 0:
                width += header_extra
            longest = max(longest, width)
        ws.column_dimensions[letter].width = min(max_width, max(min_width, longest + 1.2))


def _write_df(ws, df) -> None:
    for row in dataframe_to_rows(order_columns(df), index=False, header=True):
        ws.append(list(row))


def _paint_status_sheet(ws) -> None:
    headers = [cell.value for cell in ws[1]]
    label_idx = headers.index("判定") + 1 if "判定" in headers else None
    yen_idx = {name: headers.index(name) + 1 for name in YEN_COLUMNS if name in headers}
    wrap_names = {"理由", "原文抜粋"}
    wrap_idx = {headers.index(name) + 1 for name in wrap_names if name in headers}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        label = row[label_idx - 1].value if label_idx else None
        fill = STATUS_FILLS.get(str(label))
        for cell in row:
            cell.border = THIN
            if fill is not None:
                cell.fill = fill
            if cell.column in yen_idx.values():
                if isinstance(cell.value, float) and math.isnan(cell.value):
                    cell.value = None
                else:
                    cell.number_format = '¥#,##0'
            if cell.column in wrap_idx:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                line_count = str(cell.value or "").count("\n") + 1
                ws.row_dimensions[cell.row].height = max(18.0, min(16.0 * line_count + 8.0, 120.0))
    _style_header(ws)
    _autosize(ws)
    if "理由" in headers:
        letter = get_column_letter(headers.index("理由") + 1)
        current = ws.column_dimensions[letter].width or 42
        ws.column_dimensions[letter].width = max(12.0, min(current, 42.0) / 2)


def _write_criteria_sheet(ws) -> None:
    title_font = Font(bold=True, size=14, color=HEADER_HEX)
    body_font = Font(size=11)
    wrap = Alignment(wrap_text=True, vertical="center")
    ws.cell(1, 1, "判定の基準").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(2, 1, "抜き出したあと、プログラムで検算します。生成AIの出力をそのまま転記しません。").font = body_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

    header_row = 4
    for col, name in enumerate(("判定", "基準"), start=1):
        cell = ws.cell(header_row, col, name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN

    for offset, (name, rule) in enumerate(judgment_criteria_rows(), start=1):
        row = header_row + offset
        name_cell = ws.cell(row, 1, name)
        rule_cell = ws.cell(row, 2, rule)
        rule_cell.alignment = wrap
        fill = STATUS_FILLS["要確認"] if name.startswith("要確認") else STATUS_FILLS["転記可"]
        for cell in (name_cell, rule_cell):
            cell.font = body_font
            cell.border = THIN
            cell.fill = fill
        ws.row_dimensions[row].height = 28

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 62


def build_result_xlsx(result: ReviewResult) -> bytes:
    wb = Workbook()

    ws_all = wb.active
    ws_all.title = "転記一覧"
    _write_df(ws_all, result.rows)
    if ws_all.max_row == 1:
        ws_all.append(["（行がありません）"] + [""] * max(ws_all.max_column - 1, 0))
    _paint_status_sheet(ws_all)

    ws_issues = wb.create_sheet("要確認")
    _write_df(ws_issues, result.issues)
    if ws_issues.max_row == 1:
        ws_issues.append(["（要確認の行はありません）"] + [""] * max(ws_issues.max_column - 1, 0))
    _paint_status_sheet(ws_issues)

    ws_rule = wb.create_sheet("判定基準")
    _write_criteria_sheet(ws_rule)

    ws_meta = wb.create_sheet("処理メモ")
    meta_rows = [
        ("項目", "内容"),
        ("受取側（デモ）", CLIENT_NAME),
        ("ファイル数", result.file_count),
        ("転記可", result.ok_count),
        ("要確認", result.issue_count),
        ("転記可の税込合計", result.ok_gross),
        ("生成AIを呼んだか", "はい" if result.llm_called else "いいえ"),
        ("生成AIが項目を埋めたか", "はい" if result.used_llm else "いいえ"),
        ("処理の説明", result.source_note),
        ("鍵の扱い", "APIキーは画面または環境変数のみ。ファイルに保存しない。原文は鍵があるときだけ外部APIへ送る"),
    ]
    for row in meta_rows:
        ws_meta.append(list(row))
    _style_header(ws_meta)
    right = Alignment(horizontal="right", vertical="center")
    for row in ws_meta.iter_rows(min_row=2, min_col=1, max_col=2, max_row=ws_meta.max_row):
        label, value_cell = row[0], row[1]
        value_cell.alignment = right
        if label.value == "転記可の税込合計":
            value_cell.number_format = "¥#,##0"
        value_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws_meta.column_dimensions["A"].width = 22
    ws_meta.column_dimensions["B"].width = 56

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
